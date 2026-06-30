"""
output.py
=========
Valuation output: terminal tables + xlsx reports.

  - P/E method : print_result / write_xlsx
  - DCF method : print_dcf_result / write_dcf_xlsx   (stage 2)
  - PEG method : print_peg_result / write_peg_xlsx   (stage 2+)
"""

from __future__ import annotations
from datetime import datetime
from pathlib import Path

from . import dcf_params as P


def print_result(result) -> None:
    """Print P/E valuation summary to terminal."""
    print()
    print("=" * 64)
    print(f"  {result.ticker}  {result.name}")
    print("=" * 64)
    print(f"  Sector     : {result.sector or '?'} / {result.industry or '?'}")
    print(f"  Price      : ${result.current_price:,.2f}")
    print(f"  EPS ({result.eps_type:<8}): ${result.eps_used:,.2f}")
    print(f"  Source     : {result.source}")
    print()
    print(f"  {'Scenario':<10}{'P/E':>8}{'Target':>14}{'Return':>12}")
    print("  " + "-" * 44)
    print(f"  {'Bear':<10}{result.bear_pe:>7.1f}x ${result.bear_target:>11,.2f} {result.bear_return:>+10.1%}")
    print(f"  {'Base':<10}{result.base_pe:>7.1f}x ${result.base_target:>11,.2f} {result.base_return:>+10.1%}")
    print(f"  {'Bull':<10}{result.bull_pe:>7.1f}x ${result.bull_target:>11,.2f} {result.bull_return:>+10.1%}")
    print()

    if result.analyst_target_mean:
        print(f"  Analyst avg: ${result.analyst_target_mean:,.2f} (ref)")

    print(f"  Rationale  : {result.rationale}")

    if result.warnings:
        print()
        print("  ⚠️  Warnings:")
        for w in result.warnings:
            print(f"     - {w}")
    print()


def write_xlsx(result, output_dir: str = ".") -> str:
    """
    Write the P/E xlsx report.

    Returns:
        Output file path
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = result.ticker

    F_TITLE = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    F_HEAD = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    F_BOLD = Font(name="Arial", size=10, bold=True)
    F_BODY = Font(name="Arial", size=10)
    FILL_TITLE = PatternFill("solid", fgColor="1F3864")
    FILL_HEAD = PatternFill("solid", fgColor="5B9BD5")
    FILL_BEAR = PatternFill("solid", fgColor="FCE4D6")
    FILL_BASE = PatternFill("solid", fgColor="E2EFDA")
    FILL_BULL = PatternFill("solid", fgColor="DDEBF7")
    thin = Side(border_style="thin", color="999999")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    AL_C = Alignment(horizontal="center", vertical="center")
    AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
    AL_R = Alignment(horizontal="right", vertical="center")

    for col, w in [("A", 3), ("B", 22), ("C", 14), ("D", 14), ("E", 14), ("F", 14)]:
        ws.column_dimensions[col].width = w

    def cell(coord, val, font=None, fill=None, align=None, nf=None):
        c = ws[coord]
        c.value = val
        if font: c.font = font
        if fill: c.fill = fill
        if align: c.alignment = align
        c.border = BORDER
        if nf: c.number_format = nf

    ws.merge_cells("B2:F2")
    cell("B2", f"{result.ticker}  {result.name}", F_TITLE, FILL_TITLE, AL_C)
    ws.row_dimensions[2].height = 26

    ws.merge_cells("B3:F3")
    cell("B3", f"Generated: {datetime.now():%Y-%m-%d %H:%M} | Source: {result.source}",
         F_BODY, None, AL_C)

    cell("B5", "Price (US$)", F_BOLD, FILL_HEAD, AL_L)
    cell("C5", result.current_price, F_BODY, None, AL_R, "$#,##0.00")
    cell("B6", f"EPS ({result.eps_type})", F_BOLD, FILL_HEAD, AL_L)
    cell("C6", result.eps_used, F_BODY, None, AL_R, "$#,##0.00")
    cell("B7", "Sector", F_BOLD, FILL_HEAD, AL_L)
    ws.merge_cells("C7:F7")
    cell("C7", f"{result.sector or '?'} / {result.industry or '?'}", F_BODY, None, AL_L)

    cell("B9", "Scenario", F_HEAD, FILL_HEAD, AL_C)
    cell("C9", "Bear", F_HEAD, FILL_HEAD, AL_C)
    cell("D9", "Base", F_HEAD, FILL_HEAD, AL_C)
    cell("E9", "Bull", F_HEAD, FILL_HEAD, AL_C)
    cell("F9", "Analyst", F_HEAD, FILL_HEAD, AL_C)

    cell("B10", "P/E multiple", F_BOLD, None, AL_L)
    cell("C10", result.bear_pe, F_BODY, FILL_BEAR, AL_R, '0.0"x"')
    cell("D10", result.base_pe, F_BODY, FILL_BASE, AL_R, '0.0"x"')
    cell("E10", result.bull_pe, F_BODY, FILL_BULL, AL_R, '0.0"x"')
    cell("F10", "—", F_BODY, None, AL_C)

    cell("B11", "Target (US$)", F_BOLD, None, AL_L)
    cell("C11", result.bear_target, F_BOLD, FILL_BEAR, AL_R, "$#,##0.00")
    cell("D11", result.base_target, F_BOLD, FILL_BASE, AL_R, "$#,##0.00")
    cell("E11", result.bull_target, F_BOLD, FILL_BULL, AL_R, "$#,##0.00")
    cell("F11", result.analyst_target_mean or "—", F_BODY, None, AL_R,
         "$#,##0.00" if result.analyst_target_mean else None)

    cell("B12", "Return", F_BOLD, None, AL_L)
    cell("C12", result.bear_return, F_BODY, FILL_BEAR, AL_R, "0.0%;[Red]-0.0%")
    cell("D12", result.base_return, F_BODY, FILL_BASE, AL_R, "0.0%;[Red]-0.0%")
    cell("E12", result.bull_return, F_BODY, FILL_BULL, AL_R, "0.0%;[Red]-0.0%")
    cell("F12", "—", F_BODY, None, AL_C)

    ws.merge_cells("B14:F14")
    cell("B14", "Rationale", F_HEAD, FILL_HEAD, AL_L)
    ws.merge_cells("B15:F15")
    cell("B15", result.rationale, F_BODY, None, AL_L)
    ws.row_dimensions[15].height = 30

    if result.warnings:
        ws.merge_cells("B17:F17")
        cell("B17", "⚠️ Warnings", F_HEAD, PatternFill("solid", fgColor="C00000"), AL_L)
        ws["B17"].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        for i, w in enumerate(result.warnings):
            r = 18 + i
            ws.merge_cells(f"B{r}:F{r}")
            cell(f"B{r}", w, F_BODY, PatternFill("solid", fgColor="FFE699"), AL_L)

    last_row = 18 + len(result.warnings) + 1
    ws.merge_cells(f"B{last_row}:F{last_row}")
    cell(f"B{last_row}",
         "Generated by stock-scenario-valuation. Scenario analysis, not investment advice.",
         Font(name="Arial", size=8, italic=True, color="888888"), None, AL_L)

    ts = datetime.now().strftime("%Y%m%d")
    out_path = Path(output_dir) / f"{result.ticker}_valuation_{ts}.xlsx"
    wb.save(out_path)
    return str(out_path)


# ============================================================
# DCF (stage 2) output
# ============================================================

def _sens_header() -> str:
    base_tg = P.TERMINAL_GROWTH["base"]
    cells = "".join(f"{base_tg + d:>11.1%}" for d in P.SENSITIVITY_TG_STEPS)
    return f"    WACC\\g_inf{cells}"


def print_dcf_result(r) -> None:
    """Print DCF valuation summary to terminal."""
    w = r.wacc
    print()
    print("=" * 68)
    print(f"  {r.ticker}  {r.name}   [DCF — Discounted Cash Flow]")
    print("=" * 68)
    print(f"  Sector     : {r.sector or '?'} / {r.industry or '?'}")
    print(f"  Price      : ${r.current_price:,.2f}")
    print(f"  Base FCF   : ${r.base_fcf / 1e9:,.2f}B  ({r.fcf_source})")
    print(f"  Net debt   : ${r.net_debt / 1e9:,.2f}B")
    print(f"  WACC       : {w.wacc:.2%}  "
          f"(Re={w.cost_of_equity:.1%} / Rd={w.cost_of_debt:.1%} / "
          f"beta={w.beta_adj} / tax={w.tax_rate:.0%} / Rf={w.rf:.2%})")
    print(f"  Forecast   : {r.projection_years} years")
    print()
    print(f"  {'Scenario':<10}{'g0':>8}{'g_inf':>8}{'Exit':>8}{'Target':>14}{'Return':>12}{'TV%':>10}")
    print("  " + "-" * 72)
    for label, s in (("Bear", r.bear), ("Base", r.base), ("Bull", r.bull)):
        print(f"  {label:<10}{s.growth:>7.1%}{s.terminal_growth:>8.1%}"
              f"{s.exit_multiple:>7.0f}x ${s.target:>11,.2f} {s.ret:>+10.1%} {s.tv_share:>9.0%}")
    print()
    print(f"  Terminal   : Gordon perpetuity × {P.TERMINAL_METHOD_BLEND:.0%} + exit EV/FCF multiple "
          f"× {1 - P.TERMINAL_METHOD_BLEND:.0%}  (Base implied {r.base.implied_terminal_multiple:.0f}x)")
    print(f"  Growth     : starting growth fades to terminal over the {r.projection_years}-yr explicit period")

    if r.analyst_target_mean:
        print(f"  Analyst avg: ${r.analyst_target_mean:,.2f} (ref)")

    print()
    print("  Sensitivity (Base target/share; rows=WACC, cols=terminal growth g_inf):")
    print(_sens_header())
    for row in r.sensitivity:
        cells = "".join(
            (f"{v:>11,.2f}" if v is not None else f"{'—':>11}")
            for v in row["values"]
        )
        print(f"    {row['wacc']:>7.1%}{cells}")
    print()

    print(f"  Rationale  : {r.rationale}")

    if r.warnings:
        print()
        print("  ⚠️  Warnings:")
        for msg in r.warnings:
            print(f"     - {msg}")
    print()


def write_dcf_xlsx(r, output_dir: str = ".") -> str:
    """Write the DCF xlsx report. Returns the file path."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = f"{r.ticker}_DCF"

    F_TITLE = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    F_HEAD = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    F_BOLD = Font(name="Arial", size=10, bold=True)
    F_BODY = Font(name="Arial", size=10)
    FILL_TITLE = PatternFill("solid", fgColor="385723")     # DCF green to distinguish P/E
    FILL_HEAD = PatternFill("solid", fgColor="70AD47")
    FILL_BEAR = PatternFill("solid", fgColor="FCE4D6")
    FILL_BASE = PatternFill("solid", fgColor="E2EFDA")
    FILL_BULL = PatternFill("solid", fgColor="DDEBF7")
    thin = Side(border_style="thin", color="999999")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    AL_C = Alignment(horizontal="center", vertical="center")
    AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
    AL_R = Alignment(horizontal="right", vertical="center")

    for col, width in [("A", 3), ("B", 22), ("C", 14), ("D", 14), ("E", 14), ("F", 14)]:
        ws.column_dimensions[col].width = width

    def cell(coord, val, font=None, fill=None, align=None, nf=None):
        c = ws[coord]
        c.value = val
        if font: c.font = font
        if fill: c.fill = fill
        if align: c.alignment = align
        c.border = BORDER
        if nf: c.number_format = nf

    w = r.wacc

    ws.merge_cells("B2:F2")
    cell("B2", f"{r.ticker}  {r.name}  — DCF (Discounted Cash Flow)", F_TITLE, FILL_TITLE, AL_C)
    ws.row_dimensions[2].height = 26
    ws.merge_cells("B3:F3")
    cell("B3", f"Generated: {datetime.now():%Y-%m-%d %H:%M} | Method: FCF DCF",
         F_BODY, None, AL_C)

    cell("B5", "Price (US$)", F_BOLD, FILL_HEAD, AL_L)
    cell("C5", r.current_price, F_BODY, None, AL_R, "$#,##0.00")
    cell("B6", "Base FCF (US$)", F_BOLD, FILL_HEAD, AL_L)
    cell("C6", r.base_fcf, F_BODY, None, AL_R, "$#,##0")
    ws.merge_cells("D6:F6")
    cell("D6", r.fcf_source, F_BODY, None, AL_L)
    cell("B7", "Net debt (US$)", F_BOLD, FILL_HEAD, AL_L)
    cell("C7", r.net_debt, F_BODY, None, AL_R, "$#,##0")
    cell("B8", "WACC", F_BOLD, FILL_HEAD, AL_L)
    cell("C8", w.wacc, F_BODY, None, AL_R, "0.00%")
    ws.merge_cells("D8:F8")
    cell("D8", f"Re={w.cost_of_equity:.1%} Rd={w.cost_of_debt:.1%} "
               f"beta={w.beta_adj} tax={w.tax_rate:.0%} Rf={w.rf:.2%}", F_BODY, None, AL_L)

    cell("B10", "Scenario", F_HEAD, FILL_HEAD, AL_C)
    cell("C10", "Bear", F_HEAD, FILL_HEAD, AL_C)
    cell("D10", "Base", F_HEAD, FILL_HEAD, AL_C)
    cell("E10", "Bull", F_HEAD, FILL_HEAD, AL_C)
    cell("F10", "Analyst", F_HEAD, FILL_HEAD, AL_C)

    rows = [
        ("Starting growth g", "growth", "0.0%", None),
        ("Terminal growth g_inf", "terminal_growth", "0.0%", None),
        ("Exit EV/FCF multiple", "exit_multiple", '0.0"x"', None),
        ("Target (US$)", "target", "$#,##0.00", True),
        ("Return", "ret", "0.0%;[Red]-0.0%", None),
        ("TV % of EV", "tv_share", "0%", None),
    ]
    for i, (label, attr, nf, _bold) in enumerate(rows):
        rr = 11 + i
        cell(f"B{rr}", label, F_BOLD, None, AL_L)
        f = F_BOLD if _bold else F_BODY
        cell(f"C{rr}", getattr(r.bear, attr), f, FILL_BEAR, AL_R, nf)
        cell(f"D{rr}", getattr(r.base, attr), f, FILL_BASE, AL_R, nf)
        cell(f"E{rr}", getattr(r.bull, attr), f, FILL_BULL, AL_R, nf)
        if attr == "target":
            cell(f"F{rr}", r.analyst_target_mean or "—", F_BODY, None, AL_R,
                 "$#,##0.00" if r.analyst_target_mean else None)
        else:
            cell(f"F{rr}", "—", F_BODY, None, AL_C)

    base_tg = P.TERMINAL_GROWTH["base"]
    cell("B18", "Sensitivity: WACC \\ g_inf", F_HEAD, FILL_HEAD, AL_C)
    for j, d in enumerate(P.SENSITIVITY_TG_STEPS):
        cell(f"{chr(67 + j)}18", base_tg + d, F_HEAD, FILL_HEAD, AL_C, "0.0%")
    for i, row in enumerate(r.sensitivity):
        rr = 19 + i
        cell(f"B{rr}", row["wacc"], F_BOLD, None, AL_R, "0.0%")
        for j, v in enumerate(row["values"]):
            cell(f"{chr(67 + j)}{rr}", v if v is not None else "—",
                 F_BODY, None, AL_R, "$#,##0.00" if v is not None else None)

    base_r = 19 + len(r.sensitivity) + 1
    ws.merge_cells(f"B{base_r}:F{base_r}")
    cell(f"B{base_r}", "Rationale", F_HEAD, FILL_HEAD, AL_L)
    ws.merge_cells(f"B{base_r + 1}:F{base_r + 1}")
    cell(f"B{base_r + 1}", r.rationale, F_BODY, None, AL_L)
    ws.row_dimensions[base_r + 1].height = 40

    nxt = base_r + 2
    if r.warnings:
        ws.merge_cells(f"B{nxt}:F{nxt}")
        cell(f"B{nxt}", "⚠️ Warnings", None, PatternFill("solid", fgColor="C00000"), AL_L)
        ws[f"B{nxt}"].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        for i, msg in enumerate(r.warnings):
            rr = nxt + 1 + i
            ws.merge_cells(f"B{rr}:F{rr}")
            cell(f"B{rr}", msg, F_BODY, PatternFill("solid", fgColor="FFE699"), AL_L)
        nxt = nxt + 1 + len(r.warnings)

    ws.merge_cells(f"B{nxt + 1}:F{nxt + 1}")
    cell(f"B{nxt + 1}",
         "Generated by stock-scenario-valuation. Scenario analysis, not investment advice.",
         Font(name="Arial", size=8, italic=True, color="888888"), None, AL_L)

    ts = datetime.now().strftime("%Y%m%d")
    out_path = Path(output_dir) / f"{r.ticker}_DCF_{ts}.xlsx"
    wb.save(out_path)
    return str(out_path)


# ============================================================
# P/E vs DCF vs PEG cross-check (--method both)
# ============================================================

def print_comparison(pe, dcf, peg=None) -> None:
    """Side-by-side comparison of P/E, DCF and PEG (any may be None)."""
    ref = pe or dcf or peg
    if ref is None:
        print("  ❌ All three methods failed to value")
        return

    price = ref.current_price
    print()
    print("=" * 68)
    print(f"  {ref.ticker}  {ref.name}   [P/E vs DCF vs PEG cross-check]")
    print("=" * 68)
    print(f"  Sector     : {ref.sector or '?'} / {ref.industry or '?'}")
    print(f"  Price      : ${price:,.2f}")
    print()
    print(f"  {'Method':<10}{'Bear':>14}{'Base':>14}{'Bull':>14}{'Base ret':>12}")
    print("  " + "-" * 64)

    if pe is not None:
        print(f"  {'P/E':<9}"
              f"${pe.bear_target:>11,.2f} ${pe.base_target:>11,.2f} "
              f"${pe.bull_target:>11,.2f} {pe.base_return:>+10.1%}")
    else:
        print(f"  {'P/E':<9}{'(insufficient data, skipped)':>34}")

    if dcf is not None:
        print(f"  {'DCF':<9}"
              f"${dcf.bear.target:>11,.2f} ${dcf.base.target:>11,.2f} "
              f"${dcf.bull.target:>11,.2f} {dcf.base.ret:>+10.1%}")
    else:
        print(f"  {'DCF':<9}{'(insufficient data, skipped)':>34}")

    if peg is not None and peg.base is not None:
        tag = "PEG" if peg.applicable else "PEG*"
        print(f"  {tag:<9}"
              f"${peg.bear.target:>11,.2f} ${peg.base.target:>11,.2f} "
              f"${peg.bull.target:>11,.2f} {peg.base.ret:>+10.1%}")
        if not peg.applicable:
            print(f"  {'':<9}* PEG gating poor for this stock (low/neg growth or cyclical); targets for reference only")
    else:
        print(f"  {'PEG':<9}{'(N/A or insufficient data, skipped)':>34}")

    print()
    if peg is not None:
        print(f"  PEG ratios : trailing {peg.trailing_peg} / forward {peg.forward_peg} "
              f"(<1 cheap / 1-2 fair / >2 pricey)")
    if ref.analyst_target_mean:
        print(f"  Analyst avg: ${ref.analyst_target_mean:,.2f} (ref)")

    if pe is not None and dcf is not None and pe.base_target and dcf.base.target:
        diff = dcf.base.target / pe.base_target - 1
        print()
        print(f"  Read: DCF Base (${dcf.base.target:,.0f}) vs P/E Base "
              f"(${pe.base_target:,.0f}), diff {diff:+.0%}")
        if abs(diff) <= 0.30:
            print("        -> Close: valuation has cash-flow support, higher confidence")
        elif diff < 0:
            print("        -> DCF more conservative: market's sector-multiple premium isn't yet")
            print("           backed by cash flow; gap reflects priced-in growth (typical of growth/quality names)")
        else:
            print("        -> DCF more optimistic: cash-flow fundamentals beat the sector multiple, possibly undervalued")
    print()


# ============================================================
# PEG growth-adjusted method (third method) output
# ============================================================

def print_peg_result(r) -> None:
    """Print PEG / growth-adjusted valuation summary to terminal."""
    print()
    print("=" * 68)
    print(f"  {r.ticker}  {r.name}   [PEG — Growth-Adjusted]")
    print("=" * 68)
    print(f"  Sector     : {r.sector or '?'} / {r.industry or '?'}")
    print(f"  Price      : ${r.current_price:,.2f}")
    print()

    def g(x, pct=False):
        if x is None:
            return "—"
        return f"{x:+.1%}" if pct else f"{x}"
    print(f"  trailing : PE={g(r.trailing_pe)}  g={g(r.trailing_growth, True)}  "
          f"PEG={g(r.trailing_peg)}   (EDGAR historical EPS)")
    print(f"  forward  : PE={g(r.forward_pe)}  g={g(r.forward_growth, True)}  "
          f"PEG={g(r.forward_peg)}   (FMP estimated EPS)")
    print(f"  PEG read : <1 cheap / 1-2 fair / >2 pricey (vs its growth)")
    print()

    if r.base is not None:
        print(f"  Growth-adjusted target (growth {r.growth_used:+.1%}, {r.growth_source}, "
              f"EPS=${r.eps_used_for_target:.2f}):")
        print(f"  {'Scenario':<10}{'Tgt PEG':>9}{'Fair PE':>9}{'Target':>14}{'Return':>12}")
        print("  " + "-" * 54)
        for label, s in (("Bear", r.bear), ("Base", r.base), ("Bull", r.bull)):
            print(f"  {label:<10}{s.target_peg:>8.1f}{s.fair_pe:>8.0f}x "
                  f"${s.target:>11,.2f} {s.ret:>+10.1%}")
        print()
        if not r.applicable:
            print("  ⚠️  Note: gating poor for this stock; targets above are reference only (see warnings)")
    else:
        print("  ❌ PEG not applicable; no growth-adjusted target (see warnings below)")
    print()

    if r.analyst_target_mean:
        print(f"  Analyst avg: ${r.analyst_target_mean:,.2f} (ref)")
    print(f"  Rationale  : {r.rationale}")

    if r.warnings:
        print()
        print("  ⚠️  Warnings / applicability:")
        for w in r.warnings:
            print(f"     - {w}")
    print()


def write_peg_xlsx(r, output_dir: str = ".") -> str:
    """Write the PEG xlsx report. Returns the file path."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = f"{r.ticker}_PEG"

    F_TITLE = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    F_HEAD = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    F_BOLD = Font(name="Arial", size=10, bold=True)
    F_BODY = Font(name="Arial", size=10)
    FILL_TITLE = PatternFill("solid", fgColor="7030A0")      # PEG purple to distinguish
    FILL_HEAD = PatternFill("solid", fgColor="9E7BB5")
    FILL_BEAR = PatternFill("solid", fgColor="FCE4D6")
    FILL_BASE = PatternFill("solid", fgColor="E2EFDA")
    FILL_BULL = PatternFill("solid", fgColor="DDEBF7")
    thin = Side(border_style="thin", color="999999")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    AL_C = Alignment(horizontal="center", vertical="center")
    AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
    AL_R = Alignment(horizontal="right", vertical="center")

    for col, width in [("A", 3), ("B", 22), ("C", 14), ("D", 14), ("E", 14), ("F", 14)]:
        ws.column_dimensions[col].width = width

    def cell(coord, val, font=None, fill=None, align=None, nf=None):
        c = ws[coord]
        c.value = val
        if font: c.font = font
        if fill: c.fill = fill
        if align: c.alignment = align
        c.border = BORDER
        if nf: c.number_format = nf

    ws.merge_cells("B2:F2")
    cell("B2", f"{r.ticker}  {r.name}  — PEG (Growth-Adjusted)", F_TITLE, FILL_TITLE, AL_C)
    ws.row_dimensions[2].height = 26
    ws.merge_cells("B3:F3")
    cell("B3", f"Generated: {datetime.now():%Y-%m-%d %H:%M} | EDGAR historical + FMP estimates",
         F_BODY, None, AL_C)

    cell("B5", "Price (US$)", F_BOLD, FILL_HEAD, AL_L)
    cell("C5", r.current_price, F_BODY, None, AL_R, "$#,##0.00")
    cell("B6", "Sector", F_BOLD, FILL_HEAD, AL_L)
    ws.merge_cells("C6:F6")
    cell("C6", f"{r.sector or '?'} / {r.industry or '?'}", F_BODY, None, AL_L)

    cell("B8", "PEG ratios", F_HEAD, FILL_HEAD, AL_C)
    cell("C8", "P/E", F_HEAD, FILL_HEAD, AL_C)
    cell("D8", "EPS growth", F_HEAD, FILL_HEAD, AL_C)
    cell("E8", "PEG", F_HEAD, FILL_HEAD, AL_C)
    cell("F8", "Source", F_HEAD, FILL_HEAD, AL_C)
    cell("B9", "trailing (hist)", F_BOLD, None, AL_L)
    cell("C9", r.trailing_pe if r.trailing_pe is not None else "—", F_BODY, None, AL_R, "0.0")
    cell("D9", r.trailing_growth if r.trailing_growth is not None else "—", F_BODY, None, AL_R, "0.0%")
    cell("E9", r.trailing_peg if r.trailing_peg is not None else "—", F_BODY, None, AL_R, "0.00")
    cell("F9", "EDGAR", F_BODY, None, AL_C)
    cell("B10", "forward (est)", F_BOLD, None, AL_L)
    cell("C10", r.forward_pe if r.forward_pe is not None else "—", F_BODY, None, AL_R, "0.0")
    cell("D10", r.forward_growth if r.forward_growth is not None else "—", F_BODY, None, AL_R, "0.0%")
    cell("E10", r.forward_peg if r.forward_peg is not None else "—", F_BODY, None, AL_R, "0.00")
    cell("F10", "FMP", F_BODY, None, AL_C)

    cell("B12", "Growth-adjusted target", F_HEAD, FILL_HEAD, AL_C)
    cell("C12", "Bear", F_HEAD, FILL_HEAD, AL_C)
    cell("D12", "Base", F_HEAD, FILL_HEAD, AL_C)
    cell("E12", "Bull", F_HEAD, FILL_HEAD, AL_C)
    cell("F12", "Analyst", F_HEAD, FILL_HEAD, AL_C)
    if r.base is not None:
        rows = [("Target PEG", "target_peg", "0.0", None),
                ("Fair P/E", "fair_pe", '0"x"', None),
                ("Target (US$)", "target", "$#,##0.00", True),
                ("Return", "ret", "0.0%;[Red]-0.0%", None)]
        for i, (label, attr, nf, bold) in enumerate(rows):
            rr = 13 + i
            cell(f"B{rr}", label, F_BOLD, None, AL_L)
            f = F_BOLD if bold else F_BODY
            cell(f"C{rr}", getattr(r.bear, attr), f, FILL_BEAR, AL_R, nf)
            cell(f"D{rr}", getattr(r.base, attr), f, FILL_BASE, AL_R, nf)
            cell(f"E{rr}", getattr(r.bull, attr), f, FILL_BULL, AL_R, nf)
            if attr == "target":
                cell(f"F{rr}", r.analyst_target_mean or "—", F_BODY, None, AL_R,
                     "$#,##0.00" if r.analyst_target_mean else None)
            else:
                cell(f"F{rr}", "—", F_BODY, None, AL_C)
        nxt = 17
    else:
        ws.merge_cells("B13:F13")
        cell("B13", "PEG not applicable; no growth-adjusted target (see warnings)", F_BODY,
             PatternFill("solid", fgColor="FFE699"), AL_L)
        nxt = 14

    ws.merge_cells(f"B{nxt}:F{nxt}")
    cell(f"B{nxt}", "Rationale", F_HEAD, FILL_HEAD, AL_L)
    ws.merge_cells(f"B{nxt + 1}:F{nxt + 1}")
    cell(f"B{nxt + 1}", r.rationale, F_BODY, None, AL_L)
    ws.row_dimensions[nxt + 1].height = 40
    nxt += 2

    if r.warnings:
        ws.merge_cells(f"B{nxt}:F{nxt}")
        cell(f"B{nxt}", "⚠️ Warnings / applicability", None, PatternFill("solid", fgColor="C00000"), AL_L)
        ws[f"B{nxt}"].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        for i, msg in enumerate(r.warnings):
            rr = nxt + 1 + i
            ws.merge_cells(f"B{rr}:F{rr}")
            cell(f"B{rr}", msg, F_BODY, PatternFill("solid", fgColor="FFE699"), AL_L)
        nxt = nxt + 1 + len(r.warnings)

    ws.merge_cells(f"B{nxt + 1}:F{nxt + 1}")
    cell(f"B{nxt + 1}",
         "Generated by stock-scenario-valuation. Scenario analysis, not investment advice. "
         "PEG applies only to profitable, positive-growth stocks.",
         Font(name="Arial", size=8, italic=True, color="888888"), None, AL_L)

    ts = datetime.now().strftime("%Y%m%d")
    out_path = Path(output_dir) / f"{r.ticker}_PEG_{ts}.xlsx"
    wb.save(out_path)
    return str(out_path)
